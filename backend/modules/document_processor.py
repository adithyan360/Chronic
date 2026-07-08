import pdfplumber
import csv
import re
import json
import os
import requests
from datetime import datetime

class DocumentProcessor:
    def __init__(self):
        self.financial_keywords = {
            'revenue': ['revenue', 'total income', 'sales', 'turnover', 'total revenue', 'कुल आय', 'बिक्री'],
            'profit': ['net profit', 'pat', 'profit after tax', 'net income', 'शुद्ध लाभ'],
            'debt': ['total debt', 'borrowings', 'loans', 'debt', 'कुल ऋण', 'उधार'],
            'ebitda': ['ebitda', 'operating profit', 'परिचालन लाभ'],
            'company_name': ['company', 'name', 'ltd', 'pvt', 'limited', 'कंपनी'],
            'credit_rating': ['rating', 'credit rating', 'grade', 'crisil', 'icra', 'care']
        }
        
        # Indian-specific document patterns
        self.indian_patterns = {
            'gstr_2a': r'GSTR-2A|Form GSTR 2A',
            'gstr_3b': r'GSTR-3B|Form GSTR 3B',
            'cibil_commercial': r'CIBIL|Commercial Credit Report|TransUnion CIBIL',
            'pan': r'[A-Z]{5}[0-9]{4}[A-Z]{1}',
            'gstin': r'[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[Z]{1}[0-9A-Z]{1}',
            'cin': r'[LU][0-9]{5}[A-Z]{2}[0-9]{4}[A-Z]{3}[0-9]{6}'
        }
    
    # Add all the Indian context methods
    def classify_indian_document(self, text):
        """Classify Indian financial documents"""
        text_lower = text.lower()
        
        if re.search(self.indian_patterns['gstr_2a'], text):
            return 'GSTR-2A (Input Tax Credit)'
        elif re.search(self.indian_patterns['gstr_3b'], text):
            return 'GSTR-3B (Monthly Return)'
        elif re.search(self.indian_patterns['cibil_commercial'], text):
            return 'CIBIL Commercial Credit Report'
        elif 'annual report' in text_lower:
            return 'Annual Report'
        elif 'financial statement' in text_lower:
            return 'Financial Statement'
        else:
            return 'Unknown Document Type'
    
    def extract_indian_identifiers(self, text):
        """Extract Indian regulatory identifiers"""
        extracted = {}
        
        # Extract PAN
        pan_match = re.search(self.indian_patterns['pan'], text)
        if pan_match:
            extracted['pan'] = pan_match.group(0)
        
        # Extract GSTIN
        gstin_match = re.search(self.indian_patterns['gstin'], text)
        if gstin_match:
            extracted['gstin'] = gstin_match.group(0)
        
        # Extract CIN
        cin_match = re.search(self.indian_patterns['cin'], text)
        if cin_match:
            extracted['cin'] = cin_match.group(0)
        
        return extracted
    
    def perform_external_research(self, company_name):
        """Simulate external research for Indian companies"""
        research_data = {
            'mca_status': 'Active',
            'recent_news': [
                f'{company_name} reports strong Q4 results',
                f'Regulatory compliance update for {company_name}',
                f'{company_name} expands operations in South India'
            ],
            'regulatory_filings': [
                'Annual Return filed on time',
                'Board Resolution for loan approval',
                'Compliance certificate submitted'
            ],
            'market_sentiment': 'Positive',
            'industry_outlook': 'Stable growth expected',
            'research_timestamp': datetime.now().isoformat()
        }
        
        return research_data
    
    def process_documents(self, uploaded_files):
        """Process all uploaded documents and extract data"""
        extracted_data = {}
        
        # Process Annual Report
        if 'annual_report' in uploaded_files:
            annual_data = self.extract_from_pdf(uploaded_files['annual_report'], 'annual')
            extracted_data.update(annual_data)
        
        # Process Financial Statement
        if 'financial_statement' in uploaded_files:
            financial_data = self.extract_from_pdf(uploaded_files['financial_statement'], 'financial')
            extracted_data.update(financial_data)
        
        # Process GST Data
        if 'gst_data' in uploaded_files:
            gst_data = self.extract_from_csv(uploaded_files['gst_data'], 'gst')
            extracted_data.update(gst_data)
        
        # Process Bank Statement
        if 'bank_statement' in uploaded_files:
            bank_data = self.extract_from_csv(uploaded_files['bank_statement'], 'bank')
            extracted_data.update(bank_data)
        
        # Process Due Diligence Notes
        if 'due_diligence' in uploaded_files:
            dd_data = self.extract_from_text(uploaded_files['due_diligence'])
            extracted_data.update(dd_data)
        
        # Process Director KYC
        if 'director_kyc' in uploaded_files:
            kyc_data = self.extract_from_pdf(uploaded_files['director_kyc'], 'kyc')
            extracted_data.update(kyc_data)
        
        # Process Collateral Documents
        if 'collateral_documents' in uploaded_files:
            collateral_data = self.extract_from_pdf(uploaded_files['collateral_documents'], 'collateral')
            extracted_data.update(collateral_data)
        
        return extracted_data
    
    def extract_from_pdf(self, filepath, doc_type):
        """Extract financial data from PDF using keyword matching"""
        extracted = {}
        
        try:
            with pdfplumber.open(filepath) as pdf:
                text = ""
                for page in pdf.pages:
                    text += page.extract_text() or ""
            
            # Extract company name
            company_match = re.search(r'([A-Z][a-zA-Z\s&]+(?:Ltd|Limited|Pvt|Private|Corp|Corporation|Inc))', text)
            if company_match:
                extracted['company_name'] = company_match.group(1).strip()
            
            # Extract financial figures
            extracted.update(self.extract_financial_figures(text))
            
            # Extract credit rating
            rating_match = re.search(r'(AAA|AA\+|AA|AA\-|A\+|A|A\-|BBB\+|BBB|BBB\-|BB\+|BB|BB\-|B\+|B|B\-|CCC|CC|C|D)', text)
            if rating_match:
                extracted['credit_rating'] = rating_match.group(1)
            
            # Document type specific extraction
            if doc_type == 'kyc':
                # Extract director information
                director_match = re.search(r'director[s]?[:\s]+([a-zA-Z\s]+)', text.lower())
                if director_match:
                    extracted['director_name'] = director_match.group(1).strip().title()
                
                # Check for negative flags in KYC
                kyc_flags = ['criminal', 'fraud', 'default', 'bankruptcy']
                extracted['kyc_negative_flags'] = [flag for flag in kyc_flags if flag in text.lower()]
            
            elif doc_type == 'collateral':
                # Extract collateral information
                collateral_keywords = ['property', 'land', 'building', 'machinery', 'equipment', 'vehicle']
                extracted['collateral_types'] = [keyword for keyword in collateral_keywords if keyword in text.lower()]
                
                # Extract collateral value
                value_match = re.search(r'value[:\s]+[₹$]?([\d,]+(?:\.\d+)?)', text.lower())
                if value_match:
                    try:
                        extracted['collateral_value'] = float(value_match.group(1).replace(',', ''))
                    except ValueError:
                        pass
            
        except Exception as e:
            print(f"Error processing PDF {filepath}: {e}")
        
        return extracted
    
    def extract_financial_figures(self, text):
        """Extract numerical financial data from text"""
        extracted = {}
        
        # Common patterns for financial figures
        patterns = {
            'revenue': r'(?:revenue|total income|sales|turnover)[\s\w]*?[\s:₹$]+([\d,]+(?:\.\d+)?)',
            'profit': r'(?:net profit|pat|profit after tax)[\s\w]*?[\s:₹$]+([\d,]+(?:\.\d+)?)',
            'debt': r'(?:total debt|borrowings|loans)[\s\w]*?[\s:₹$]+([\d,]+(?:\.\d+)?)',
            'ebitda_margin': r'(?:ebitda margin|operating margin)[\s\w]*?[\s:]+([\d.]+)%?'
        }
        
        for key, pattern in patterns.items():
            match = re.search(pattern, text.lower())
            if match:
                value_str = match.group(1).replace(',', '')
                try:
                    if key == 'ebitda_margin':
                        extracted[key] = float(value_str)
                    else:
                        # Convert to lakhs/crores if needed
                        value = float(value_str)
                        if 'crore' in text.lower():
                            value *= 10000000
                        elif 'lakh' in text.lower():
                            value *= 100000
                        extracted[key] = int(value)
                except ValueError:
                    pass
        
        return extracted
    
    def extract_from_csv(self, filepath, data_type):
        """Extract data from CSV/Excel files"""
        extracted = {}
        
        try:
            # Handle Excel files
            if filepath.endswith(('.xlsx', '.xls')):
                try:
                    if filepath.endswith('.xlsx'):
                        import openpyxl
                        wb = openpyxl.load_workbook(filepath)
                        ws = wb.active
                        rows = []
                        headers = [cell.value for cell in ws[1]]
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
                            rows.append(row_dict)
                    else:  # .xls file
                        import xlrd
                        wb = xlrd.open_workbook(filepath)
                        ws = wb.sheet_by_index(0)
                        headers = [ws.cell_value(0, col) for col in range(ws.ncols)]
                        rows = []
                        for row_idx in range(1, ws.nrows):
                            row_dict = {headers[col]: ws.cell_value(row_idx, col) for col in range(ws.ncols)}
                            rows.append(row_dict)
                except ImportError:
                    # Fallback: skip Excel processing
                    print(f"Excel libraries not available for {filepath}")
                    return extracted
                except Exception as e:
                    print(f"Excel processing failed for {filepath}: {e}")
                    return extracted
            else:
                # Handle CSV files with multiple encodings
                for encoding in ['utf-8', 'latin-1', 'cp1252']:
                    try:
                        with open(filepath, 'r', encoding=encoding) as f:
                            reader = csv.DictReader(f)
                            rows = list(reader)
                        break
                    except UnicodeDecodeError:
                        continue
                else:
                    print(f"Could not decode CSV file {filepath}")
                    return extracted
            
            if data_type == 'gst':
                # GST data analysis
                sales_total = 0
                amount_total = 0
                
                for row in rows:
                    # Try different column name variations
                    for sales_col in ['sales', 'Sales', 'SALES', 'amount', 'Amount']:
                        if sales_col in row and row[sales_col]:
                            try:
                                sales_total += float(row[sales_col].replace(',', ''))
                                break
                            except (ValueError, AttributeError):
                                pass
                
                extracted['gst_sales'] = sales_total
                extracted['gst_total'] = amount_total or sales_total
            
            elif data_type == 'bank':
                # Bank statement analysis
                credit_total = 0
                debit_total = 0
                
                for row in rows:
                    # Credit/deposits
                    for credit_col in ['credit', 'Credit', 'CREDIT', 'deposits', 'Deposits']:
                        if credit_col in row and row[credit_col]:
                            try:
                                credit_total += float(row[credit_col].replace(',', ''))
                                break
                            except (ValueError, AttributeError):
                                pass
                    
                    # Debit/withdrawals
                    for debit_col in ['debit', 'Debit', 'DEBIT', 'withdrawals', 'Withdrawals']:
                        if debit_col in row and row[debit_col]:
                            try:
                                debit_total += float(row[debit_col].replace(',', ''))
                                break
                            except (ValueError, AttributeError):
                                pass
                
                extracted['bank_deposits'] = credit_total
                extracted['bank_withdrawals'] = debit_total
                
                # Detect circular transactions
                extracted['circular_transactions'] = self.detect_circular_transactions_simple(rows)
        
        except Exception as e:
            print(f"Error processing CSV {filepath}: {e}")
        
        return extracted
    
    def detect_circular_transactions_simple(self, rows):
        """Simple circular transaction detection without pandas"""
        circular_count = 0
        amounts = []
        
        try:
            for row in rows:
                # Try to find amount column
                for amount_col in ['amount', 'Amount', 'AMOUNT', 'credit', 'Credit', 'debit', 'Debit']:
                    if amount_col in row and row[amount_col]:
                        try:
                            amount = float(row[amount_col].replace(',', ''))
                            amounts.append(amount)
                            break
                        except (ValueError, AttributeError):
                            pass
            
            # Count duplicate amounts (simple heuristic)
            amount_counts = {}
            for amount in amounts:
                amount_counts[amount] = amount_counts.get(amount, 0) + 1
            
            # Count amounts that appear more than twice
            circular_count = sum(1 for count in amount_counts.values() if count > 2)
        
        except Exception:
            pass
        
        return circular_count
    
    def extract_from_text(self, filepath):
        """Extract data from text files (due diligence notes)"""
        extracted = {}
        
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                text = f.read().lower()
            
            # Check for negative keywords
            negative_keywords = ['litigation', 'lawsuit', 'fraud', 'default', 'loss', 'decline']
            extracted['negative_flags'] = [word for word in negative_keywords if word in text]
            
            # Check for positive keywords
            positive_keywords = ['collateral', 'security', 'guarantee', 'asset']
            extracted['positive_flags'] = [word for word in positive_keywords if word in text]
        
        except Exception as e:
            print(f"Error processing text file {filepath}: {e}")
        
        return extracted